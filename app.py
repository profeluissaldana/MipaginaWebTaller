from flask import Flask, render_template, request, redirect, url_for, session, flash, Response, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date
from sqlalchemy import extract
import csv
import io
import os

# Imports específicos para reportes avanzados
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

app = Flask(__name__)
app.secret_key = 'clave_secreta_para_el_taller_tecnico'

# =========================================================================
# CONFIGURACIÓN DINÁMICA DE LA BASE DE DATOS (LOCAL VS PRODUCCIÓN)
# =========================================================================
URL_BASE_DATOS = os.environ.get('DATABASE_URL')

if URL_BASE_DATOS:
    # Si estamos en Render, usamos PostgreSQL
    # Corrección técnica de compatibilidad de hilos en strings de SQLAlchemy:
    if URL_BASE_DATOS.startswith("postgres://"):
        URL_BASE_DATOS = URL_BASE_DATOS.replace("postgres://", "postgresql://", 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = URL_BASE_DATOS
else:
    # Si estamos trabajando en modo local, sigue usando SQLite
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///taller_computacion.db'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)


# =========================================================================
# MODELOS DE LA BASE DE DATOS (ORM)
# =========================================================================

class Turno(db.Model):
    __tablename__ = 'turnos'
    id = db.Column(db.Integer, primary_key=True)
    nombre_grupo = db.Column(db.String(100), nullable=False) 
    turno_horario = db.Column(db.String(20), nullable=False)  
    dias_cursada = db.Column(db.String(100), nullable=False)  

    alumnos = db.relationship('Alumno', backref='turno_assigned', lazy=True)
    clases_dadas = db.relationship('Clase', backref='turno_clase', lazy=True)

    def __repr__(self):
        return f'<Turno {self.nombre_grupo}>'


class Alumno(db.Model):
    __tablename__ = 'alumnos'
    id = db.Column(db.Integer, primary_key=True)
    usuario = db.Column(db.String(50), unique=True, nullable=False)
    contrasena = db.Column(db.String(255), nullable=False) 
    email = db.Column(db.String(120), unique=True, nullable=False)
    dni = db.Column(db.String(20), unique=True, nullable=False)
    apellido = db.Column(db.String(100), nullable=False)
    nombre = db.Column(db.String(100), nullable=False)
    taller = db.Column(db.String(100), default="Taller de Computación")
    observaciones_generales = db.Column(db.Text, nullable=True) 
    
    turno_id = db.Column(db.Integer, db.ForeignKey('turnos.id'), nullable=False)

    asistencias = db.relationship('Asistencia', backref='alumno', lazy=True)
    actions_auditoria = db.relationship('Auditoria', backref='alumno', lazy=True)

    def __repr__(self):
        return f'<Alumno {self.apellido}, {self.nombre}>'


class Clase(db.Model):
    __tablename__ = 'clases'
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, default=datetime.today, nullable=False)
    tema_dado = db.Column(db.String(255), nullable=False) 
    detalle_clase = db.Column(db.Text, nullable=True)     
    
    turno_id = db.Column(db.Integer, db.ForeignKey('turnos.id'), nullable=False)

    def __repr__(self):
        return f'<Clase Fecha: {self.fecha} - Tema: {self.tema_dado}>'


class Asistencia(db.Model):
    __tablename__ = 'asistencias'
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, default=datetime.today, nullable=False)
    trimestre = db.Column(db.Integer, nullable=False) 
    hora_entrada = db.Column(db.DateTime, default=datetime.now, nullable=False)
    hora_salida = db.Column(db.DateTime, nullable=True) 
    estado = db.Column(db.String(1), default='P', nullable=False) # P, A, T, J
    comentario_diario = db.Column(db.Text, nullable=True) 
    
    alumno_id = db.Column(db.Integer, db.ForeignKey('alumnos.id'), nullable=False)

    @property
    def tiempo_permanencia(self):
        if self.hora_salida:
            diferencia = self.hora_salida - self.hora_entrada
            minutos_totales = int(diferencia.total_seconds() / 60)
            horas = minutos_totales // 60
            minutos = minutes_totales = minutos_totales % 60
            return f"{horas}h {minutos}m"
        return "En el taller"

    def __repr__(self):
        return f'<Asistencia Alumno_ID: {self.alumno_id} - Estado: {self.estado}>'


class Auditoria(db.Model):
    __tablename__ = 'auditorias'
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, default=datetime.now, nullable=False)
    accion = db.Column(db.String(255), nullable=False)      
    url_visitada = db.Column(db.String(255), nullable=False)    
    direccion_ip = db.Column(db.String(50), nullable=True) 
    
    alumno_id = db.Column(db.Integer, db.ForeignKey('alumnos.id'), nullable=False)

    def __repr__(self):
        return f'<Auditoria Alumno_ID {self.alumno_id} - {self.accion}>'


# =========================================================================
# FUNCIÓN ANALÍTICA DE ASISTENCIA (Cómputo de Media Falta)
# =========================================================================
def generar_reporte_asistencia(alumno_id, anio_actual=2026):
    asistencias = Asistencia.query.filter(
        Asistencia.alumno_id == alumno_id,
        extract('year', Asistencia.fecha) == anio_actual
    ).all()
    
    reporte = {
        'anual': {'totales': 0, 'asistencias': 0, 'inasistencias': 0, 'porc_asist': 0, 'porc_inasist': 0},
        'trimestres': {
            1: {'totales': 0, 'asistencias': 0, 'inasistencias': 0, 'porc_asist': 0, 'porc_inasist': 0},
            2: {'totales': 0, 'asistencias': 0, 'inasistencias': 0, 'porc_asist': 0, 'porc_inasist': 0},
            3: {'totales': 0, 'asistencias': 0, 'inasistencias': 0, 'porc_asist': 0, 'porc_inasist': 0}
        },
        'mensual': {}
    }
    
    if not asistencias:
        return reporte
        
    for asis in asistencias:
        mes = asis.fecha.month
        trimestre = asis.trimestre
        
        if mes not in reporte['mensual']:
            reporte['mensual'][mes] = {'totales': 0, 'asistencias': 0, 'inasistencias': 0, 'porc_asist': 0, 'porc_inasist': 0}
            
        if asis.estado == 'P':
            val_asist = 1.0
            val_inasist = 0.0
        elif asis.estado == 'T':
            val_asist = 0.5    
            val_inasist = 0.5  # Llegada Tarde computa media falta
        elif asis.estado in ['A', 'J']:
            val_asist = 0.0
            val_inasist = 1.0  

        reporte['mensual'][mes]['totales'] += 1
        reporte['trimestres'][trimestre]['totales'] += 1
        reporte['anual']['totales'] += 1

        reporte['mensual'][mes]['asistencias'] += val_asist
        reporte['mensual'][mes]['inasistencias'] += val_inasist
        reporte['trimestres'][trimestre]['asistencias'] += val_asist
        reporte['trimestres'][trimestre]['inasistencias'] += val_inasist
        reporte['anual']['asistencias'] += val_asist
        reporte['anual']['inasistencias'] += val_inasist

    for m in reporte['mensual']:
        tot = reporte['mensual'][m]['totales']
        if tot > 0:
            reporte['mensual'][m]['porc_asist'] = round((reporte['mensual'][m]['asistencias'] / tot) * 100, 1)
            reporte['mensual'][m]['porc_inasist'] = round((reporte['mensual'][m]['inasistencias'] / tot) * 100, 1)
            
    for t in reporte['trimestres']:
        tot = reporte['trimestres'][t]['totales']
        if tot > 0:
            reporte['trimestres'][t]['porc_asist'] = round((reporte['trimestres'][t]['asistencias'] / tot) * 100, 1)
            reporte['trimestres'][t]['porc_inasist'] = round((reporte['trimestres'][t]['inasistencias'] / tot) * 100, 1)
            
    tot_anual = reporte['anual']['totales']
    if tot_anual > 0:
        reporte['anual']['porc_asist'] = round((reporte['anual']['asistencias'] / tot_anual) * 100, 1)
        reporte['anual']['porc_inasist'] = round((reporte['anual']['inasistencias'] / tot_anual) * 100, 1)

    return reporte


# =========================================================================
# RUTAS DE AUTENTICACIÓN (LOGIN / LOGOUT)
# =========================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'usuario_id' in session:
        return redirect(url_for('inicio'))
        
    if request.method == 'POST':
        user_input = request.form.get('usuario')
        pass_input = request.form.get('contrasena')
        
        alumno = Alumno.query.filter_by(usuario=user_input).first()
        
        if alumno and alumno.contrasena == pass_input:
            session['usuario_id'] = alumno.id
            session['usuario_nombre'] = f"{alumno.nombre} {alumno.apellido}"
            
            # CONTROL DE ACCESO: Definimos si es el docente/admin
            if alumno.usuario.strip().lower() == 'luis':
                session['is_admin'] = True
            else:
                 session['is_admin'] = False
            
            nueva_accion = Auditoria(accion="Inició sesión en la plataforma", url_visitada="/login", alumno_id=alumno.id)
            db.session.add(nueva_accion)
            db.session.commit()
            
            flash('¡Bienvenido al sistema del Taller!', 'success')
            return redirect(url_for('inicio'))
        else:
            flash('Usuario o contraseña incorrectos.', 'danger')
            
    return render_template('login.html')


@app.route('/logout')
def logout():
    if 'usuario_id' in session:
        nueva_accion = Auditoria(accion="Cerró sesión", url_visitada="/logout", alumno_id=session['usuario_id'])
        db.session.add(nueva_accion)
        db.session.commit()
    session.clear()
    flash('Sesión cerrada correctamente.', 'info')
    return redirect(url_for('login'))


# =========================================================================
# PANEL DOCENTE: DESCARGAR PLANTILLA CSV MODELO (INTEGRACIÓN 6 TURNOS)
# =========================================================================
@app.route('/docente/descargar_plantilla_csv')
def descargar_plantilla_csv():
    if not session.get('is_admin'):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('login'))
        
    output = io.StringIO()
    columnas = ['usuario', 'contrasena', 'email', 'dni', 'apellido', 'nombre', 'taller', 'turno_id', 'observaciones_generales']
    
    escritor = csv.writer(output, delimiter=';')
    escritor.writerow(columnas)
    
    # Filas de ejemplo actualizadas para reflejar la distribución detallada (ej. Turno 1, 3 y 5)
    escritor.writerow(['amartinez', 'alumno2026', 'ana.martinez@tecnica.edu.ar', '46123456', 'Martínez', 'Ana', 'Taller de Computación', '1', '5to A - Mañana Lun/Mie'])
    escritor.writerow(['frossi', 'clave5to', 'facu.rossi@tecnica.edu.ar', '47987654', 'Rossi', 'Facundo', 'Taller de Computación', '3', '5to C - Tarde Martes'])
    escritor.writerow(['llopez', 'clave789', 'laura.l@tecnica.edu.ar', '48222333', 'López', 'Laura', 'Taller de Computación', '5', '5to C - Tarde Viernes'])
    
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-disposition": "attachment; filename=plantilla_alumnos.csv"}
    )

# =========================================================================
# PANEL DOCENTE: VISUALIZAR GRUPOS Y ALUMNOS CARGADOS
# =========================================================================
@app.route('/docente/ver_grupos')
def ver_grupos():
    if not session.get('is_admin'):
        flash('Acceso denegado. Solo el docente puede ver el desglose de grupos.', 'danger')
        return redirect(url_for('login'))
        
    # Traemos todos los turnos de la base de datos
    turnos = Turno.query.all()
    
    # Capturamos si el docente hizo clic para ver los alumnos de un turno específico
    turno_ver_id = request.args.get('turno_id', type=int)
    alumnos_del_turno = []
    turno_seleccionado = None
    
    if turno_ver_id:
        turno_seleccionado = Turno.query.get(turno_ver_id)
        if turno_seleccionado:
            alumnos_del_turno = Alumno.query.filter_by(turno_id=turno_ver_id).order_by(Alumno.apellido.asc()).all()

    return render_template('ver_grupos.html', turnos=turnos, alumnos=alumnos_del_turno, turno_sel=turno_seleccionado)

# =========================================================================
# PANEL DOCENTE: ALTA DE ALUMNOS (FORMULARIO MANUAL + EXCEL CSV VALIDADO)
# =========================================================================
@app.route('/docente/cargar_alumno', methods=['GET', 'POST'])
def cargar_alumno():
    if not session.get('is_admin'):
        flash('Acceso denegado. Solo el docente puede registrar alumnos.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        # --- SUBIDA POR ARCHIVO EXCEL/CSV ---
        if 'archivo_csv' in request.files and request.files['archivo_csv'].filename != '':
            file = request.files['archivo_csv']
            if not file.filename.endswith('.csv'):
                flash('Error: El archivo debe ser .csv', 'danger')
                return redirect(url_for('cargar_alumno'))
            
            try:
                # CAMBIO OPERATIVO: Usamos latin-1 para que Excel procese los acentos sin arrojar error 0xed
                stream = io.StringIO(file.stream.read().decode("latin-1"), newline=None)
                lector_csv = csv.DictReader(stream, delimiter=';')
                
                columnas_requeridas = ['usuario', 'contrasena', 'email', 'dni', 'apellido', 'nombre', 'turno_id']
                if not all(col in lector_csv.fieldnames for col in columnas_requeridas):
                    flash('Error: Formato de CSV incorrecto en las columnas.', 'danger')
                    return redirect(url_for('cargar_alumno'))
                
                creados, errores = 0, 0
                for fila in lector_csv:
                    user_input = fila['usuario'].strip().lower()
                    dni_input = fila['dni'].strip()
                    email_input = fila['email'].strip()
                    
                    # Validación y limpieza estricta del nuevo rango de turnos (1 al 6)
                    try:
                        csv_turno_id = int(fila['turno_id'].strip())
                        if csv_turno_id < 1 or csv_turno_id > 6:
                            errores += 1
                            continue  # Omite si el ID de turno está fuera del rango válido escolar
                    except ValueError:
                        errores += 1
                        continue # Omite si no es un número válido
                    
                    if Alumno.query.filter_by(usuario=user_input).first() or Alumno.query.filter_by(dni=dni_input).first():
                        errores += 1
                        continue
                        
                    nuevo = Alumno(
                        usuario=user_input, contrasena=fila['contrasena'].strip(), email=email_input,
                        dni=dni_input, apellido=fila['apellido'].strip(), nombre=fila['nombre'].strip(),
                        taller=fila.get('taller', 'Taller de Computación').strip() or "Taller de Computación",
                        turno_id=csv_turno_id,
                        observaciones_generales=fila.get('observaciones_generales', '').strip() or None
                    )
                    db.session.add(nuevo)
                    creados += 1
                
                db.session.commit()
                flash(f'📥 Importación exitosa. Creados: {creados}. Duplicados/Errores de Turno: {errores}.', 'success')
                return redirect(url_for('cargar_alumno'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error al procesar el archivo CSV: {str(e)}', 'danger')
                return redirect(url_for('cargar_alumno'))

        # --- CARGA POR FORMULARIO MANUAL ---
        else:
            user_input = request.form.get('usuario').strip().lower()
            dni_input = request.form.get('dni').strip()
            form_turno_id = request.form.get('turno_id', type=int)
            
            # Validación de rango de turno para resguardar la consistencia manual
            if form_turno_id < 1 or form_turno_id > 6:
                flash('Error: El turno seleccionado no pertenece a un esquema válido (1-6).', 'danger')
                return redirect(url_for('cargar_alumno'))
            
            if Alumno.query.filter_by(usuario=user_input).first() or Alumno.query.filter_by(dni=dni_input).first():
                flash('Error: El usuario o DNI ya existen.', 'danger')
            else:
                nuevo = Alumno(
                    usuario=user_input, contrasena=request.form.get('contrasena'),
                    email=request.form.get('email').strip(), dni=dni_input,
                    apellido=request.form.get('apellido').strip(), nombre=request.form.get('nombre').strip(),
                    taller=request.form.get('taller') or "Taller de Computación",
                    turno_id=form_turno_id,
                    observaciones_generales=request.form.get('observaciones_generales') or None
                )
                db.session.add(nuevo)
                db.session.commit()
                flash(f'¡Alumno {nuevo.apellido} registrado correctamente!', 'success')
                return redirect(url_for('cargar_alumno'))

    turnos = Turno.query.all()
    return render_template('cargar_alumno.html', turnos=turnos)


# =========================================================================
# PANEL DOCENTE: TOMAR ASISTENCIA Y LIBRO DE TEMAS
# =========================================================================
@app.route('/docente/asistencia', methods=['GET', 'POST'])
def tomar_asistencia():
    if not session.get('is_admin'):
        flash('Acceso denegado. Se requieren permisos de docente.', 'danger')
        return redirect(url_for('login'))

    turnos = Turno.query.all()
    turno_seleccionado_id = request.args.get('turno_id', type=int)
    
    # CAPTURAMOS EL SUBGRUPO DESDE EL DESPLEGABLE GENERAL (Ej: "Grupo 3")
    subgrupo_rotacion = request.args.get('subgrupo_rotacion', type=str)
    
    alumnos_grupo = []
    
    if turno_seleccionado_id:
        # SI ES EL TURNO 6 Y SELECCIONASTE UN GRUPO ROTATIVO
        if turno_seleccionado_id == 6 and subgrupo_rotacion:
            alumnos_grupo = Alumno.query.filter(
                Alumno.turno_id == 6,
                Alumno.observaciones_generales.like(f"%{subgrupo_rotacion}%")
            ).all()
        else:
            # Para los turnos normales (1 al 5) trae a todos directo
            alumnos_grupo = Alumno.query.filter_by(turno_id=turno_seleccionado_id).all()

    if request.method == 'POST':
        turno_id = request.form.get('turno_id', type=int)
        subgrupo_post = request.form.get('subgrupo_rotacion_post', '')
        
        # En el Libro de Temas dejamos asentado automáticamente qué grupo rotó hoy
        tema_original = request.form.get('tema_dado')
        if turno_id == 6 and subgrupo_post:
            tema_final = f"[{subgrupo_post}] - {tema_original}"
        else:
            tema_final = tema_original

        nueva_clase = Clase(
            tema_dado=tema_final,
            detalle_clase=request.form.get('detalle_clase'),
            turno_id=turno_id  # Guarda automáticamente Turno 6 para el jueves
        )
        db.session.add(nueva_clase)
        
        # Procesamos la asistencia SOLO de los alumnos del subgrupo que trajimos a pantalla
        for alu in alumnos_grupo:
            estado_alu = request.form.get(f'asistencia_{alu.id}', 'P')
            comentario_alu = request.form.get(f'comentario_{alu.id}', '').strip()
            
            # Si es jueves, le dejamos una marca automática en su asistencia individual
            if turno_id == 6 and subgrupo_post:
                comentario_final = f"[{subgrupo_post}] {comentario_alu}".strip()
            else:
                comentario_final = comentario_alu if comentario_alu else None

            asistencia_jornada = Asistencia(
                trimestre=request.form.get('trimestre', type=int),
                estado=estado_alu,
                comentario_diario=comentario_final,
                alumno_id=alu.id
            )
            db.session.add(asistencia_jornada)
            
        db.session.commit()
        flash('¡Asistencia de la rotación guardada con éxito!', 'success')
        return redirect(url_for('inicio'))

    return render_template(
        'tomar_asistencia.html', 
        turnos=turnos, 
        alumnos=alumnos_grupo, 
        turno_sel_id=turno_seleccionado_id,
        subgrupo_sel=subgrupo_rotacion
    )


# =========================================================================
# PANEL DOCENTE: CENTRO DE REPORTES Y FILTROS
# =========================================================================
@app.route('/docente/reportes')
def panel_reportes():
    if not session.get('is_admin'):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('login'))
    turnos = Turno.query.all()
    return render_template('reportes.html', turnos=turnos)


# =========================================================================
# PANEL DOCENTE: EXPORTAR ASISTENCIAS A EXCEL NATIVO (.XLSX)
# =========================================================================
@app.route('/docente/exportar/excel')
def exportar_excel():
    if not session.get('is_admin'):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('login'))

    turno_id = request.args.get('turno_id', type=int)
    trimestre = request.args.get('trimestre', type=int)

    query = db.session.query(Asistencia).join(Alumno)
    if turno_id:
        query = query.filter(Alumno.turno_id == turno_id)
    if trimestre:
        query = query.filter(Asistencia.trimestre == trimestre)
    
    registros = query.order_by(Asistencia.fecha.desc(), Alumno.apellido.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Asistencias"

    fuente_titulo = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    fuente_cabecera = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    fuente_datos = Font(name="Arial", size=10)
    
    fill_titulo = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_cabecera = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    
    borde_fino = Border(
        left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
    )

    ws.merge_cells('A1:G1')
    ws['A1'] = "REPORTE CONSOLIDADO DE ASISTENCIAS - TALLER DE COMPUTACIÓN"
    ws['A1'].font = fuente_titulo
    ws['A1'].fill = fill_titulo
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    headers = ["Fecha", "Trimestre", "Apellido", "Nombre", "DNI", "Estado", "Observaciones"]
    ws.append([]) 
    ws.append(headers) 

    for col_num, header in enumerate(headers, 1):
        celda = ws.cell(row=3, column=col_num)
        celda.font = fuente_cabecera
        celda.fill = fill_cabecera
        celda.alignment = Alignment(horizontal="center")

    mapa_estados = {'P': 'Presente', 'A': 'Ausente', 'T': 'Tarde', 'J': 'Justificado'}

    for r in registros:
        estado_desc = mapa_estados.get(r.estado, r.estado)
        fila = [
            r.fecha.strftime('%d/%m/%Y'),
            f"{r.trimestre}° Trim.",
            r.alumno.apellido,
            r.alumno.nombre,
            r.alumno.dni,
            estado_desc,
            r.comentario_diario or ""
        ]
        ws.append(fila)
        
        for col_num in range(1, 8):
            celda = ws.cell(row=ws.max_row, column=col_num)
            celda.font = fuente_datos
            celda.border = borde_fino
            if col_num in [1, 2, 5, 6]: 
                celda.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name="Reporte_Asistencias.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# =========================================================================
# PANEL DOCENTE: EXPORTAR ASISTENCIAS A PDF FORMATEADO
# =========================================================================
@app.route('/docente/exportar/pdf')
def exportar_pdf():
    if not session.get('is_admin'):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('login'))

    turno_id = request.args.get('turno_id', type=int)
    trimestre = request.args.get('trimestre', type=int)

    query = db.session.query(Asistencia).join(Alumno)
    if turno_id:
        query = query.filter(Alumno.turno_id == turno_id)
    if trimestre:
        query = query.filter(Asistencia.trimestre == trimestre)
    
    registros = query.order_by(Asistencia.fecha.desc(), Alumno.apellido.asc()).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    styles = getSampleStyleSheet()
    style_titulo = ParagraphStyle(name='TituloDoc', parent=styles['Heading1'], fontSize=16, leading=20, textColor=colors.HexColor("#1F4E78"), alignment=1, spaceAfter=15)
    style_sub = ParagraphStyle(name='SubDoc', parent=styles['Normal'], fontSize=10, leading=14, textColor=colors.HexColor("#555555"), alignment=1, spaceAfter=20)
    style_celda = ParagraphStyle(name='CeldaTabla', parent=styles['Normal'], fontSize=9, leading=11)
    style_cb_tabla = ParagraphStyle(name='CabTabla', parent=styles['Normal'], fontSize=9, leading=11, fontName="Helvetica-Bold", textColor=colors.white)

    story.append(Paragraph("<b>REGISTRO OFICIAL DE ASISTENCIAS</b>", style_titulo))
    story.append(Paragraph(f"Reporte emitido el: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Taller de Computación", style_sub))
    story.append(Spacer(1, 10))

    tabla_datos = [[
        Paragraph("Fecha", style_cb_tabla),
        Paragraph("Trim.", style_cb_tabla),
        Paragraph("Alumno", style_cb_tabla),
        Paragraph("DNI", style_cb_tabla),
        Paragraph("Est.", style_cb_tabla),
        Paragraph("Observaciones", style_cb_tabla)
    ]]

    mapa_estados = {'P': 'Presente', 'A': 'Ausente', 'T': 'Tarde', 'J': 'Justificado'}

    for r in registros:
        tabla_datos.append([
            Paragraph(r.fecha.strftime('%d/%m/%Y'), style_celda),
            Paragraph(f"{r.trimestre}°", style_celda),
            Paragraph(f"{r.alumno.apellido}, {r.alumno.nombre}", style_celda),
            Paragraph(r.alumno.dni, style_celda),
            Paragraph(mapa_estados.get(r.estado, r.estado), style_celda),
            Paragraph(r.comentario_diario or "", style_celda)
        ])

    t = Table(tabla_datos, colWidths=[65, 35, 140, 65, 60, 185])
    
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#2C3E50")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D0D0D0")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F9F9F9")]),
        ('TOPPADDING', (0,1), (-1,-1), 5),
        ('BOTTOMPADDING', (0,1), (-1,-1), 5),
    ]))
    
    story.append(t)
    doc.build(story)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="Reporte_Asistencias.pdf",
        mimetype="application/pdf"
    )


# =========================================================================
# FUNCIÓN DE CONTROL DE ACCESO (PROTECCIÓN DE VISTAS)
# =========================================================================
def usuario_autenticado():
    """Devuelve True si el usuario inició sesión, de lo contrario redirige al login"""
    return 'usuario_id' in session

# =========================================================================
# PANEL DOCENTE: CONSULTA HISTÓRICA DE ASISTENCIAS POR GRUPO
# =========================================================================
# =========================================================================
# PANEL DOCENTE: CONSULTA HISTÓRICA DE ASISTENCIAS CON FECHA POR DEFECTO
# =========================================================================
# =========================================================================
# PANEL DOCENTE: CONSULTA HISTÓRICA DE ASISTENCIAS (CORREGIDO Y OPTIMIZADO)
# =========================================================================
@app.route('/docente/historial_asistencias')
def historial_asistencias():
    if not session.get('is_admin'):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('login'))
    
    turnos = Turno.query.all()
    
    # Capturamos los filtros limpiando espacios o valores vacíos del formulario
    turno_id_raw = request.args.get('turno_id', '')
    fecha_filtro = request.args.get('fecha', type=str)
    
    # Procesamos el turno de forma segura (si es vacío o 'None', no se filtra)
    turno_id = None
    if turno_id_raw and turno_id_raw != 'None' and str(turno_id_raw).strip() != '':
        try:
            turno_id = int(turno_id_raw)
        except ValueError:
            turno_id = None
    
    # Si el calendario vino vacío, tomamos la fecha de hoy por defecto (YYYY-MM-DD)
    if not fecha_filtro or fecha_filtro.strip() == '':
        fecha_filtro = date.today().strftime('%Y-%m-%d')
    
    # Iniciamos la consulta base uniendo la asistencia con el alumno
    query = db.session.query(Asistencia).join(Alumno)
    
    # Aplicamos el filtro de turno SOLO si se seleccionó una comisión válida
    if turno_id:
        query = query.filter(Alumno.turno_id == turno_id)
        
    if fecha_filtro:
        try:
            # Convertimos el string del input HTML a un objeto date real de Python
            fecha_dt = datetime.strptime(fecha_filtro, '%Y-%m-%d').date()
            # Filtramos exactamente por ese objeto date
            query = query.filter(Asistencia.fecha == fecha_dt)
        except ValueError:
            # Si el formato fallara por alguna razón, no rompemos la app
            pass

    # Traemos los resultados ordenados alfabéticamente por apellido
    asistencias = query.order_by(Alumno.apellido.asc()).all()
    
    return render_template('historial_asistencias.html', 
                           turnos=turnos, 
                           asistencias=asistencias, 
                           turno_sel_id=turno_id, 
                           fecha_sel=fecha_filtro)


# =========================================================================
# PANEL DOCENTE: CORREGIR/EDITAR EL ESTADO DE UNA ASISTENCIA YA GRABADA
# =========================================================================
@app.route('/docente/corregir_asistencia/<int:id>', methods=['POST'])
def corregir_asistencia(id):
    if not session.get('is_admin'):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('login'))
        
    asistencia = Asistencia.query.get_or_404(id)
    
    # Capturamos las modificaciones del formulario modal
    nuevo_estado = request.form.get('nuevo_estado')
    nuevo_comentario = request.form.get('nuevo_comentario', '').strip()
    
    # Parámetros de retorno para mantener la misma vista filtrada al recargar
    turno_retorno = request.form.get('turno_retorno', '')
    fecha_retorno = request.form.get('fecha_retorno', '')
    
    try:
        asistencia.estado = nuevo_estado
        asistencia.comentario_diario = nuevo_comentario if nuevo_comentario else None
        db.session.commit()
        flash(f'Asistencia de {asistencia.alumno.apellido} corregida con éxito.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al corregir la asistencia: {str(e)}', 'danger')
        
    # Aseguramos que retorne mandando explícitamente los filtros limpios
    return redirect(url_for('historial_asistencias', turno_id=turno_retorno, fecha=fecha_retorno))


# =========================================================================
# PANEL DOCENTE: MODIFICAR DATOS DE UN ALUMNO
# =========================================================================
@app.route('/docente/editar_alumno/<int:id>', methods=['GET', 'POST'])
def editar_alumno(id):
    if not session.get('is_admin'):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('login'))
        
    alumno = Alumno.query.get_or_404(id)
    turnos = Turno.query.all()
    
    if request.method == 'POST':
        alumno.nombre = request.form.get('nombre').strip()
        alumno.apellido = request.form.get('apellido').strip()
        alumno.dni = request.form.get('dni').strip()
        alumno.email = request.form.get('email').strip()
        alumno.turno_id = request.form.get('turno_id', type=int)
        alumno.observaciones_generales = request.form.get('observaciones_generales').strip() or None
        
        try:
            db.session.commit()
            flash(f'Datos de {alumno.apellido} actualizados con éxito.', 'success')
            return redirect(url_for('ver_grupos', turno_id=alumno.turno_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el alumno: El DNI o usuario ya existen.', 'danger')
            
    return render_template('editar_alumno.html', alumno=alumno, turnos=turnos)


# =========================================================================
# PANEL DOCENTE: ELIMINAR UN ALUMNO (Y SU HISTORIAL POR CASCADA)
# =========================================================================
@app.route('/docente/eliminar_alumno/<int:id>', methods=['POST'])
def eliminar_alumno(id):
    if not session.get('is_admin'):
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('login'))
        
    alumno = Alumno.query.get_or_404(id)
    turno_previo = alumno.turno_id
    
    try:
        # 1. Eliminamos registros asociados en cascada para evitar errores de clave foránea
        Asistencia.query.filter_by(alumno_id=alumno.id).delete()
        Auditoria.query.filter_by(alumno_id=alumno.id).delete()
        
        # 2. Eliminamos al alumno de la base de datos
        db.session.delete(alumno)
        db.session.commit()
        
        flash(f'El alumno {alumno.apellido}, {alumno.nombre} fue removido del sistema correctamente.', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'No se pudo eliminar al alumno: {str(e)}', 'danger')
        
    return redirect(url_for('ver_grupos', turno_id=turno_previo))


# =========================================================================
# RUTA RAÍZ PRINCIPAL / INICIO (AGREGADA PARA EVITAR EL ERROR 404)
# =========================================================================
@app.route('/')
def inicio():
    if not usuario_autenticado():
        return redirect(url_for('login'))
        
    # Agregamos esto para que la pantalla de inicio pueda dibujar las comisiones
    turnos = Turno.query.all()
    return render_template('inicio.html', turnos=turnos)


# =========================================================================
# RUTAS DE CONTENIDOS ESTÁTICOS INFORMATIVOS (PROTEGIDAS)
# =========================================================================

@app.route('/Ofimatica')
def ofimatica(): 
    if not usuario_autenticado(): return redirect(url_for('login'))
    return render_template('ofimatica.html')

@app.route('/VisualStudioCode')
def visualStudioCode(): 
    if not usuario_autenticado(): return redirect(url_for('login'))
    return render_template('visualstudiocode.html')

@app.route('/HTML_CSS')
def htmlCss(): 
    if not usuario_autenticado(): return redirect(url_for('login'))
    return render_template('htmlcss.html')

@app.route('/Javascript')
def javaScript(): 
    if not usuario_autenticado(): return redirect(url_for('login'))
    return render_template('javascript.html')

@app.route('/Git_Github')
def gitGitHub(): 
    if not usuario_autenticado(): return redirect(url_for('login'))
    return render_template('gitgithub.html')

@app.route('/arquitectura_flask')
def arquitecturaFlask(): 
    if not usuario_autenticado(): return redirect(url_for('login'))
    return render_template('arquitecturaflask.html')

@app.route('/frameworkflask')
def frameworkFlask(): 
    if not usuario_autenticado(): return redirect(url_for('login'))
    return render_template('frameworkflask.html')

@app.route('/arduino_uno')
def arduinoUno(): 
    if not usuario_autenticado(): return redirect(url_for('login'))
    return render_template('arduino_uno.html')

@app.route('/servidorrender')
def servidorRender(): 
    if not usuario_autenticado(): return redirect(url_for('login'))
    return render_template('servidorrender.html')


# =========================================================================
# PANEL DOCENTE: REINICIO DE CICLO LECTIVO (BORRADO SEGURO DESDE EL PANEL)
# =========================================================================
@app.route('/docente/reiniciar_ciclo', methods=['GET', 'POST'])
def reiniciar_ciclo():
    if not session.get('is_admin'):
        flash('Acceso denegado. Solo el administrador puede realizar esta acción.', 'danger')
        return redirect(url_for('login'))
        
    if request.method == 'POST':
        confirmacion = request.form.get('confirmacion', '').strip().lower()
        
        if confirmacion == 'reiniciar2026':
            try:
                docente_actual = Alumno.query.filter_by(usuario='luis').first()
                
                # Resguardo seguro de tus credenciales
                if docente_actual:
                    datos_luis = {
                        'usuario': docente_actual.usuario,
                        'contrasena': docente_actual.contrasena,
                        'email': docente_actual.email,
                        'dni': docente_actual.dni,
                        'apellido': docente_actual.apellido,
                        'nombre': docente_actual.nombre
                    }
                else:
                    datos_luis = {
                        'usuario': 'luis',
                        'contrasena': 'admin2026', # Tu clave por defecto si no existe registro previo
                        'email': 'luis.saldana@tecnica.edu.ar',
                        'dni': '12345678',
                        'apellido': 'Saldaña',
                        'nombre': 'Luis'
                    }

                # Vaciado completo de tablas
                db.drop_all()
                db.create_all()
                
                # Recreación automatizada de los 6 turnos oficiales de la institución
                inicializar_turnos_oficiales()

                # Reestablecemos tu usuario de forma inmediata
                admin_luis = Alumno(
                    usuario=datos_luis['usuario'],
                    contrasena=datos_luis['contrasena'],
                    email=datos_luis['email'],
                    dni=datos_luis['dni'],
                    apellido=datos_luis['apellido'],
                    nombre=datos_luis['nombre'],
                    taller='Taller de Computación',
                    turno_id=1
                )
                db.session.add(admin_luis)
                db.session.commit()
                
                flash('¡El sistema se ha reiniciado con éxito! Estructura limpia y cuenta de docente preservada.', 'success')
                return redirect(url_for('inicio'))
                
            except Exception as e:
                db.session.rollback()
                flash(f'Ocurrió un error al reiniciar el sistema: {str(e)}', 'danger')
        else:
            flash('La palabra de confirmación es incorrecta. No se realizaron cambios.', 'danger')
            
    return render_template('reiniciar_ciclo.html')


# =========================================================================
# VISTA DEL RENDIMIENTO DEL ALUMNO (PANEL INDIVIDUAL)
# =========================================================================
@app.route('/alumno/<int:id>/rendimiento')
def rendimiento_alumno(id):
    alumno = Alumno.query.get_or_404(id)
    datos_asistencia = generar_reporte_asistencia(alumno.id)
    meses_nombres = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio", 
                     7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}
    return render_template('rendimiento.html', alumno=alumno, reporte=datos_asistencia, meses=meses_nombres)


# =========================================================================
# CONFIGURACIÓN E INICIALIZACIÓN DE DATASET (INTEGRACIÓN 6 TURNOS)
# =========================================================================
def inicializar_turnos_oficiales():
    """Chequea y monta los 6 turnos oficiales de la E.E.T.P. N° 614"""
    turnos_objetivo = [
        {"id": 1, "nombre": "Mañana - Grupo A (Lun/Mie)", "horario": "7:30 - 11:10", "dias": "Lunes y Miércoles"},
        {"id": 2, "nombre": "Mañana - Grupo B (Mar/Vie)", "horario": "7:30 - 11:10", "dias": "Martes y Viernes"},
        {"id": 3, "nombre": "Tarde - Martes", "horario": "13:10 - 16:50", "dias": "Martes"},
        {"id": 4, "nombre": "Tarde - Miércoles", "horario": "13:10 - 16:50", "dias": "Miércoles"},
        {"id": 5, "nombre": "Tarde - Viernes", "horario": "13:10 - 16:50", "dias": "Viernes"},
        {"id": 6, "nombre": "Tarde - Jueves Rotativo", "horario": "13:10 - 16:50", "dias": "Jueves"}
    ]
    
    for t_obj in turnos_objetivo:
        turno_existente = Turno.query.get(t_obj["id"])
        if not turno_existente:
            nuevo_turno = Turno(
                id=t_obj["id"],
                nombre_grupo=t_obj["nombre"],
                turno_horario=t_obj["horario"],
                dias_cursada=t_obj["dias"]
            )
            db.session.add(nuevo_turno)
        else:
            turno_existente.nombre_grupo = t_obj["nombre"]
            turno_existente.turno_horario = t_obj["horario"]
            turno_existente.dias_cursada = t_obj["dias"]
            
    db.session.commit()

# Crear tablas y sembrar datos de configuración inicial dentro del contexto de la app
with app.app_context():
    db.create_all()
    inicializar_turnos_oficiales()

if __name__ == '__main__':
    app.run(debug=True)